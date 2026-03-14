"""Excel Tool - Read and manipulate Excel files (.xlsx, .xlsm)."""
import os
from datetime import datetime
from typing import Any
from fastmcp import FastMCP
from ..file_system_toolkits.security import get_secure_path

def register_tools(mcp: FastMCP) -> None:
    """Register Excel tools with the MCP server."""
    @mcp.tool()
    def excel_read(
        path: str,
        workspace_id: str,
        agent_id: str,
        session_id: str,
        sheet: str | None = None,
        limit: int | None = None,
        offset: int = 0,
    ) -> dict:
        """
        Read an Excel file and return its contents.
        Args:
            path: Path to the Excel file (relative to session sandbox)
            workspace_id: Workspace identifier
            agent_id: Agent identifier
            session_id: Session identifier
            sheet: Sheet name to read (default: active sheet)
            limit: Maximum number of rows to return (None = all rows)
            offset: Number of rows to skip from the beginning (after header)
        Returns:
            dict with success status, data, and metadata
        """
        if offset < 0 or (limit is not None and limit < 0):
            return {"error": "offset and limit must be non-negative"}
        try:
            from openpyxl import load_workbook
        except ImportError:
            return {
                "error": (
                    "openpyxl not installed. Install with: "
                    "pip install openpyxl or pip install tools[excel]"
                )
            }
        try:
            secure_path = get_secure_path(path, workspace_id, agent_id, session_id)
            if not os.path.exists(secure_path):
                return {"error": f"File not found: {path}"}
            if not path.lower().endswith((".xlsx", ".xlsm")):
                return {"error": "File must have .xlsx or .xlsm extension"}
            # Load workbook in read-only mode for better performance
            wb = load_workbook(secure_path, read_only=True, data_only=True)
            try:
                # Get the specified sheet or active sheet
                if sheet:
                    if sheet not in wb.sheetnames:
                        return {
                            "error": f"Sheet '{sheet}' not found. Available sheets: {wb.sheetnames}"
                        }
                    ws = wb[sheet]
                else:
                    ws = wb.active
                if ws is None:
                    return {"error": "Workbook has no active sheet"}
                # Read all rows
                all_rows = []
                for row in ws.iter_rows(values_only=True):
                    # Convert cell values to serializable format
                    converted_row = [_convert_cell_value(cell) for cell in row]
                    all_rows.append(converted_row)
                if not all_rows:
                    return {
                        "success": True,
                        "path": path,
                        "sheet_name": ws.title,
                        "columns": [],
                        "column_count": 0,
                        "rows": [],
                        "row_count": 0,
                        "total_rows": 0,
                        "offset": offset,
                        "limit": limit,
                    }
                # First row as headers
                columns = all_rows[0] if all_rows else []
                data_rows = all_rows[1:] # Rows without header
                # Apply offset and limit to data rows
                total_rows = len(data_rows)
                if offset > 0:
                    data_rows = data_rows[offset:]
                if limit is not None:
                    data_rows = data_rows[:limit]
                # Convert rows to list of dicts with column names as keys
                rows_as_dicts = []
                for row in data_rows:
                    row_dict = {}
                    for i, value in enumerate(row):
                        if i < len(columns) and columns[i]:
                            col_name = columns[i]
                        else:
                            col_name = f"Column_{i + 1}"
                        row_dict[str(col_name)] = value
                    rows_as_dicts.append(row_dict)
                # Format column names
                formatted_columns = [
                    str(c) if c is not None else f"Column_{i + 1}" for i, c in enumerate(columns)
                ]
                return {
                    "success": True,
                    "path": path,
                    "sheet_name": ws.title,
                    "columns": formatted_columns,
                    "column_count": len(columns),
                    "rows": rows_as_dicts,
                    "row_count": len(rows_as_dicts),
                    "total_rows": total_rows,
                    "offset": offset,
                    "limit": limit,
                }
            finally:
                wb.close()
        except Exception as e:
            return {"error": f"Failed to read Excel file: {str(e)}"}

    @mcp.tool()
    def excel_write(
        path: str,
        workspace_id: str,
        agent_id: str,
        session_id: str,
        columns: list[str],
        rows: list[dict],
        sheet: str = "Sheet1",
    ) -> dict:
        """
        Write data to a new Excel file.
        Args:
            path: Path to the Excel file (relative to session sandbox)
            workspace_id: Workspace identifier
            agent_id: Agent identifier
            session_id: Session identifier
            columns: List of column names for the header
            rows: List of dictionaries, each representing a row
            sheet: Name for the sheet (default: "Sheet1")
        Returns:
            dict with success status and metadata
        """
        try:
            from openpyxl import Workbook
        except ImportError:
            return {
                "error": (
                    "openpyxl not installed. Install with: "
                    "pip install openpyxl or pip install tools[excel]"
                )
            }
        try:
            secure_path = get_secure_path(path, workspace_id, agent_id, session_id)
            if not path.lower().endswith((".xlsx", ".xlsm")):
                return {"error": "File must have .xlsx or .xlsm extension"}
            if not columns:
                return {"error": "columns cannot be empty"}
            # Create parent directories if needed
            parent_dir = os.path.dirname(secure_path)
            if parent_dir:
                os.makedirs(parent_dir, exist_ok=True)
            # Create new workbook
            wb = Workbook()
            ws = wb.active
            if ws is None:
                return {"error": "Failed to create worksheet"}
            ws.title = sheet
            # Write header row
            for col_idx, col_name in enumerate(columns, start=1):
                ws.cell(row=1, column=col_idx, value=col_name)
            # Write data rows
            for row_idx, row_data in enumerate(rows, start=2):
                for col_idx, col_name in enumerate(columns, start=1):
                    value = row_data.get(col_name, "")
                    ws.cell(row=row_idx, column=col_idx, value=value)
            # Save workbook
            wb.save(secure_path)
            wb.close()
            return {
                "success": True,
                "path": path,
                "sheet_name": sheet,
                "columns": columns,
                "column_count": len(columns),
                "rows_written": len(rows),
            }
        except Exception as e:
            return {"error": f"Failed to write Excel file: {str(e)}"}

    @mcp.tool()
    def excel_append(
        path: str,
        workspace_id: str,
        agent_id: str,
        session_id: str,
        rows: list[dict],
        sheet: str | None = None,
    ) -> dict:
        """
        Append rows to an existing Excel file.
        Args:
            path: Path to the Excel file (relative to session sandbox)
            workspace_id: Workspace identifier
            agent_id: Agent identifier
            session_id: Session identifier
            rows: List of dictionaries to append, keys should match existing columns
            sheet: Sheet name to append to (default: active sheet)
        Returns:
            dict with success status and metadata
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            return {
                "error": (
                    "openpyxl not installed. Install with: "
                    "pip install openpyxl or pip install tools[excel]"
                )
            }
        try:
            secure_path = get_secure_path(path, workspace_id, agent_id, session_id)
            if not os.path.exists(secure_path):
                return {"error": f"File not found: {path}. Use excel_write to create a new file."}
            if not path.lower().endswith((".xlsx", ".xlsm")):
                return {"error": "File must have .xlsx or .xlsm extension"}
            if not rows:
                return {"error": "rows cannot be empty"}
            # Load existing workbook
            wb = load_workbook(secure_path)
            try:
                # Get the specified sheet or active sheet
                if sheet:
                    if sheet not in wb.sheetnames:
                        return {
                            "error": (
                                f"Sheet '{sheet}' not found. Available sheets: {wb.sheetnames}"
                            )
                        }
                    ws = wb[sheet]
                else:
                    ws = wb.active
                if ws is None:
                    return {"error": "Workbook has no active sheet"}
                # Get existing columns from first row
                columns = []
                for cell in ws[1]:
                    columns.append(str(cell.value) if cell.value is not None else "")
                if not columns or all(c == "" for c in columns):
                    return {"error": "Excel file has no headers in the first row"}
                # Find the next empty row
                next_row = ws.max_row + 1
                # Append rows
                for row_data in rows:
                    for col_idx, col_name in enumerate(columns, start=1):
                        value = row_data.get(col_name, "")
                        ws.cell(row=next_row, column=col_idx, value=value)
                    next_row += 1
                # Save workbook
                wb.save(secure_path)
                # Get new total row count (excluding header)
                total_rows = next_row - 2 # -1 for header, -1 because next_row was incremented
                return {
                    "success": True,
                    "path": path,
                    "sheet_name": ws.title,
                    "rows_appended": len(rows),
                    "total_rows": total_rows,
                }
            finally:
                wb.close()
        except Exception as e:
            return {"error": f"Failed to append to Excel file: {str(e)}"}

    @mcp.tool()
    def excel_info(
        path: str,
        workspace_id: str,
        agent_id: str,
        session_id: str,
    ) -> dict:
        """
        Get metadata about an Excel file without reading all data.
        Args:
            path: Path to the Excel file (relative to session sandbox)
            workspace_id: Workspace identifier
            agent_id: Agent identifier
            session_id: Session identifier
        Returns:
            dict with file metadata (sheets, columns per sheet, row counts, file size)
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            return {
                "error": (
                    "openpyxl not installed. Install with: "
                    "pip install openpyxl or pip install tools[excel]"
                )
            }
        try:
            secure_path = get_secure_path(path, workspace_id, agent_id, session_id)
            if not os.path.exists(secure_path):
                return {"error": f"File not found: {path}"}
            if not path.lower().endswith((".xlsx", ".xlsm")):
                return {"error": "File must have .xlsx or .xlsm extension"}
            # Get file size
            file_size = os.path.getsize(secure_path)
            # Load workbook in read-only mode
            wb = load_workbook(secure_path, read_only=True, data_only=True)
            try:
                sheets_info = []
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    # Get columns from first row
                    columns = []
                    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
                    if first_row:
                        columns = [
                            str(c) if c is not None else f"Column_{i + 1}"
                            for i, c in enumerate(first_row)
                        ]
                    # Count rows (excluding header)
                    row_count = 0
                    for _ in ws.iter_rows(min_row=2, values_only=True):
                        row_count += 1
                    sheets_info.append(
                        {
                            "name": sheet_name,
                            "columns": columns,
                            "column_count": len(columns),
                            "row_count": row_count,
                        }
                    )
                return {
                    "success": True,
                    "path": path,
                    "file_size_bytes": file_size,
                    "sheet_count": len(wb.sheetnames),
                    "sheet_names": wb.sheetnames,
                    "sheets": sheets_info,
                }
            finally:
                wb.close()
        except Exception as e:
            return {"error": f"Failed to get Excel info: {str(e)}"}

    @mcp.tool()
    def excel_sheet_list(
        path: str,
        workspace_id: str,
        agent_id: str,
        session_id: str,
    ) -> dict:
        """
        List all sheet names in an Excel file.
        Args:
            path: Path to the Excel file (relative to session sandbox)
            workspace_id: Workspace identifier
            agent_id: Agent identifier
            session_id: Session identifier
        Returns:
            dict with list of sheet names
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            return {
                "error": (
                    "openpyxl not installed. Install with: "
                    "pip install openpyxl or pip install tools[excel]"
                )
            }
        try:
            secure_path = get_secure_path(path, workspace_id, agent_id, session_id)
            if not os.path.exists(secure_path):
                return {"error": f"File not found: {path}"}
            if not path.lower().endswith((".xlsx", ".xlsm")):
                return {"error": "File must have .xlsx or .xlsm extension"}
            # Load workbook in read-only mode (minimal memory usage)
            wb = load_workbook(secure_path, read_only=True)
            try:
                return {
                    "success": True,
                    "path": path,
                    "sheet_names": wb.sheetnames,
                    "sheet_count": len(wb.sheetnames),
                }
            finally:
                wb.close()
        except Exception as e:
            return {"error": f"Failed to list sheets: {str(e)}"}

    @mcp.tool()
    def excel_sql(
        path: str,
        workspace_id: str,
        agent_id: str,
        session_id: str,
        query: str,
        sheet: str | None = None,
    ) -> dict:
        """
        Query an Excel file using SQL (powered by DuckDB).
        Each sheet is available as a table with its sheet name (spaces replaced
        with underscores). Use 'data' as alias for the specified/active sheet.
        Args:
            path: Path to the Excel file (relative to session sandbox)
            workspace_id: Workspace identifier
            agent_id: Agent identifier
            session_id: Session identifier
            query: SQL query. Use 'data' for the target sheet, or sheet names
                   (with spaces as underscores) to query/join multiple sheets.
            sheet: Sheet to use as 'data' table (default: first sheet)
        Returns:
            dict with query results, columns, and row count
        Examples:
            # Simple query on default sheet
            query="SELECT * FROM data WHERE price > 100"
            # Aggregate data
            query="SELECT category, SUM(amount) as total FROM data GROUP BY category"
            # Join multiple sheets (sheet names: 'Sales', 'Products')
            query="SELECT s.*, p.name FROM Sales s JOIN Products p ON s.product_id = p.id"
        """
        try:
            import duckdb
        except ImportError:
            return {
                "error": (
                    "DuckDB not installed. Install with: "
                    "pip install duckdb or pip install tools[sql]"
                )
            }
        try:
            from openpyxl import load_workbook
        except ImportError:
            return {
                "error": (
                    "openpyxl not installed. Install with: "
                    "pip install openpyxl or pip install tools[excel]"
                )
            }
        try:
            secure_path = get_secure_path(path, workspace_id, agent_id, session_id)
            if not os.path.exists(secure_path):
                return {"error": f"File not found: {path}"}
            if not path.lower().endswith((".xlsx", ".xlsm")):
                return {"error": "File must have .xlsx or .xlsm extension"}
            if not query or not query.strip():
                return {"error": "query cannot be empty"}
            # Security: only allow SELECT statements
            query_upper = query.strip().upper()
            if not query_upper.startswith("SELECT"):
                return {"error": "Only SELECT queries are allowed for security reasons"}
            # Disallowed keywords
            disallowed = [
                "INSERT",
                "UPDATE",
                "DELETE",
                "DROP",
                "CREATE",
                "ALTER",
                "TRUNCATE",
                "EXEC",
                "EXECUTE",
            ]
            for keyword in disallowed:
                if keyword in query_upper:
                    return {"error": f"'{keyword}' is not allowed in queries"}
            # Load workbook
            wb = load_workbook(secure_path, read_only=True, data_only=True)
            try:
                # Determine target sheet for 'data' alias
                if sheet:
                    if sheet not in wb.sheetnames:
                        return {"error": (f"Sheet '{sheet}' not found. Available: {wb.sheetnames}")}
                    target_sheet = sheet
                else:
                    target_sheet = wb.sheetnames[0]
                # Load all sheets into DuckDB
                import pandas as pd
                con = duckdb.connect(":memory:")
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    rows = list(ws.iter_rows(values_only=True))
                    if not rows:
                        continue
                    # Headers from first row
                    headers = [
                        str(c) if c is not None else f"Column_{i + 1}"
                        for i, c in enumerate(rows[0])
                    ]
                    # Data rows
                    records = []
                    for row in rows[1:]:
                        record = {}
                        for i, val in enumerate(row):
                            col = headers[i] if i < len(headers) else f"Column_{i + 1}"
                            record[col] = _convert_cell_value(val)
                        records.append(record)
                    # Create table (sanitize name: spaces -> underscores)
                    table_name = sheet_name.replace(" ", "_").replace("-", "_")
                    if records:
                        df = pd.DataFrame(records)
                        con.register(f"temp_{table_name}", df)
                        con.execute(
                            f'CREATE TABLE "{table_name}" AS SELECT * FROM temp_{table_name}'
                        )
                    else:
                        # Empty table
                        cols_sql = ", ".join(f'"{h}" VARCHAR' for h in headers)
                        con.execute(f'CREATE TABLE "{table_name}" ({cols_sql})')
                    # Create 'data' alias for target sheet
                    if sheet_name == target_sheet:
                        con.execute(f'CREATE VIEW data AS SELECT * FROM "{table_name}"')
                all_sheet_names = list(wb.sheetnames)
            finally:
                wb.close()
            # Execute query (workbook already closed, only DuckDB needed)
            try:
                result = con.execute(query)
                columns = [desc[0] for desc in result.description]
                rows = result.fetchall()
            finally:
                con.close()
            # Convert to dicts
            rows_as_dicts = [dict(zip(columns, row)) for row in rows]
            return {
                "success": True,
                "path": path,
                "target_sheet": target_sheet,
                "available_sheets": all_sheet_names,
                "query": query,
                "columns": columns,
                "column_count": len(columns),
                "rows": rows_as_dicts,
                "row_count": len(rows_as_dicts),
            }
        except Exception as e:
            error_msg = str(e)
            if "Catalog Error" in error_msg or "Table" in error_msg:
                return {
                    "error": f"SQL error: {error_msg}. "
                    "Use 'data' for target sheet or sheet names with underscores."
                }
            return {"error": f"Query failed: {error_msg}"}

    @mcp.tool()
    def excel_search(
        path: str,
        workspace_id: str,
        agent_id: str,
        session_id: str,
        search_term: str,
        sheet: str | None = None,
        case_sensitive: bool = False,
        match_type: str = "contains",
    ) -> dict:
        """
        Search for values across Excel sheets.
        Args:
            path: Path to the Excel file (relative to session sandbox)
            workspace_id: Workspace identifier
            agent_id: Agent identifier
            session_id: Session identifier
            search_term: Text to search for
            sheet: Specific sheet to search (default: search all sheets)
            case_sensitive: Whether search is case-sensitive (default: False)
            match_type: 'contains', 'exact', 'starts_with', or 'ends_with'
        Returns:
            dict with list of matches containing sheet, row, column, and value
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            return {
                "error": (
                    "openpyxl not installed. Install with: "
                    "pip install openpyxl or pip install tools[excel]"
                )
            }
        try:
            secure_path = get_secure_path(path, workspace_id, agent_id, session_id)
            if not os.path.exists(secure_path):
                return {"error": f"File not found: {path}"}
            if not path.lower().endswith((".xlsx", ".xlsm")):
                return {"error": "File must have .xlsx or .xlsm extension"}
            if not search_term:
                return {"error": "search_term cannot be empty"}
            if match_type not in ("contains", "exact", "starts_with", "ends_with"):
                return {
                    "error": "match_type must be 'contains', 'exact', 'starts_with', or 'ends_with'"
                }
            # Prepare search term
            term = search_term if case_sensitive else search_term.lower()
            # Load workbook
            wb = load_workbook(secure_path, read_only=True, data_only=True)
            try:
                sheets_to_search = [sheet] if sheet else wb.sheetnames
                if sheet and sheet not in wb.sheetnames:
                    return {"error": f"Sheet '{sheet}' not found. Available: {wb.sheetnames}"}
                matches = []
                for sheet_name in sheets_to_search:
                    ws = wb[sheet_name]
                    # Get headers for column names
                    headers = []
                    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
                    if first_row:
                        headers = [
                            str(c) if c is not None else f"Column_{i + 1}"
                            for i, c in enumerate(first_row)
                        ]
                    # Search data rows only (skip header row)
                    for row_idx, row in enumerate(
                        ws.iter_rows(min_row=2, values_only=True), start=2
                    ):
                        for col_idx, cell_value in enumerate(row):
                            if cell_value is None:
                                continue
                            # Convert to string for comparison
                            cell_str = str(cell_value)
                            compare_val = cell_str if case_sensitive else cell_str.lower()
                            # Check match
                            is_match = False
                            if match_type == "contains":
                                is_match = term in compare_val
                            elif match_type == "exact":
                                is_match = term == compare_val
                            elif match_type == "starts_with":
                                is_match = compare_val.startswith(term)
                            elif match_type == "ends_with":
                                is_match = compare_val.endswith(term)
                            if is_match:
                                col_name = (
                                    headers[col_idx]
                                    if col_idx < len(headers)
                                    else f"Column_{col_idx + 1}"
                                )
                                matches.append(
                                    {
                                        "sheet": sheet_name,
                                        "row": row_idx,
                                        "column": col_name,
                                        "column_index": col_idx + 1,
                                        "value": _convert_cell_value(cell_value),
                                    }
                                )
                return {
                    "success": True,
                    "path": path,
                    "search_term": search_term,
                    "match_type": match_type,
                    "case_sensitive": case_sensitive,
                    "sheets_searched": sheets_to_search,
                    "matches": matches,
                    "match_count": len(matches),
                }
            finally:
                wb.close()
        except Exception as e:
            return {"error": f"Search failed: {str(e)}"}

    @mcp.tool()
    def excel_add_sheet(
        path: str,
        workspace_id: str,
        agent_id: str,
        session_id: str,
        sheet_name: str = "NewSheet",
    ) -> dict:
        """
        Add a new sheet to an existing Excel file.
        Args:
            path: Path to the Excel file (relative to session sandbox)
            workspace_id: Workspace identifier
            agent_id: Agent identifier
            session_id: Session identifier
            sheet_name: Name for the new sheet (default: "NewSheet")
        Returns:
            dict with success status and metadata
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            return {
                "error": (
                    "openpyxl not installed. Install with: "
                    "pip install openpyxl or pip install tools[excel]"
                )
            }
        try:
            secure_path = get_secure_path(path, workspace_id, agent_id, session_id)
            if not os.path.exists(secure_path):
                return {"error": f"File not found: {path}"}
            if not path.lower().endswith((".xlsx", ".xlsm")):
                return {"error": "File must have .xlsx or .xlsm extension"}
            if not sheet_name:
                return {"error": "sheet_name cannot be empty"}
            wb = load_workbook(secure_path)
            try:
                if sheet_name in wb.sheetnames:
                    return {"error": f"Sheet '{sheet_name}' already exists"}
                wb.create_sheet(sheet_name)
                wb.save(secure_path)
                return {
                    "success": True,
                    "path": path,
                    "new_sheet_name": sheet_name,
                    "sheet_count": len(wb.sheetnames),
                }
            finally:
                wb.close()
        except Exception as e:
            return {"error": f"Failed to add sheet: {str(e)}"}

    @mcp.tool()
    def excel_rename_sheet(
        path: str,
        workspace_id: str,
        agent_id: str,
        session_id: str,
        old_name: str,
        new_name: str,
    ) -> dict:
        """
        Rename a sheet in an Excel file.
        Args:
            path: Path to the Excel file (relative to session sandbox)
            workspace_id: Workspace identifier
            agent_id: Agent identifier
            session_id: Session identifier
            old_name: Current sheet name
            new_name: New sheet name
        Returns:
            dict with success status and metadata
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            return {
                "error": (
                    "openpyxl not installed. Install with: "
                    "pip install openpyxl or pip install tools[excel]"
                )
            }
        try:
            secure_path = get_secure_path(path, workspace_id, agent_id, session_id)
            if not os.path.exists(secure_path):
                return {"error": f"File not found: {path}"}
            if not path.lower().endswith((".xlsx", ".xlsm")):
                return {"error": "File must have .xlsx or .xlsm extension"}
            if not old_name or not new_name:
                return {"error": "old_name and new_name cannot be empty"}
            wb = load_workbook(secure_path)
            try:
                if old_name not in wb.sheetnames:
                    return {"error": f"Sheet '{old_name}' not found"}
                if new_name in wb.sheetnames:
                    return {"error": f"Sheet '{new_name}' already exists"}
                ws = wb[old_name]
                ws.title = new_name
                wb.save(secure_path)
                return {
                    "success": True,
                    "path": path,
                    "old_name": old_name,
                    "new_name": new_name,
                }
            finally:
                wb.close()
        except Exception as e:
            return {"error": f"Failed to rename sheet: {str(e)}"}

    @mcp.tool()
    def excel_delete_sheet(
        path: str,
        workspace_id: str,
        agent_id: str,
        session_id: str,
        sheet_name: str,
    ) -> dict:
        """
        Delete a sheet from an Excel file.
        Args:
            path: Path to the Excel file (relative to session sandbox)
            workspace_id: Workspace identifier
            agent_id: Agent identifier
            session_id: Session identifier
            sheet_name: Name of the sheet to delete
        Returns:
            dict with success status and metadata
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            return {
                "error": (
                    "openpyxl not installed. Install with: "
                    "pip install openpyxl or pip install tools[excel]"
                )
            }
        try:
            secure_path = get_secure_path(path, workspace_id, agent_id, session_id)
            if not os.path.exists(secure_path):
                return {"error": f"File not found: {path}"}
            if not path.lower().endswith((".xlsx", ".xlsm")):
                return {"error": "File must have .xlsx or .xlsm extension"}
            if not sheet_name:
                return {"error": "sheet_name cannot be empty"}
            wb = load_workbook(secure_path)
            try:
                if sheet_name not in wb.sheetnames:
                    return {"error": f"Sheet '{sheet_name}' not found"}
                if len(wb.sheetnames) == 1:
                    return {"error": "Cannot delete the last sheet in the workbook"}
                del wb[sheet_name]
                wb.save(secure_path)
                return {
                    "success": True,
                    "path": path,
                    "deleted_sheet": sheet_name,
                    "sheet_count": len(wb.sheetnames),
                }
            finally:
                wb.close()
        except Exception as e:
            return {"error": f"Failed to delete sheet: {str(e)}"}

    @mcp.tool()
    def excel_update_cell(
        path: str,
        workspace_id: str,
        agent_id: str,
        session_id: str,
        sheet: str | None = None,
        row: int,
        column: str | int,
        value: Any,
    ) -> dict:
        """
        Update a specific cell in an Excel sheet.
        Args:
            path: Path to the Excel file (relative to session sandbox)
            workspace_id: Workspace identifier
            agent_id: Agent identifier
            session_id: Session identifier
            sheet: Sheet name (default: active sheet)
            row: 1-indexed row number
            column: Column name (str) or 1-indexed column index (int)
            value: New value for the cell
        Returns:
            dict with success status and metadata
        """
        try:
            from openpyxl import load_workbook
            from openpyxl.utils import get_column_letter
        except ImportError:
            return {
                "error": (
                    "openpyxl not installed. Install with: "
                    "pip install openpyxl or pip install tools[excel]"
                )
            }
        try:
            secure_path = get_secure_path(path, workspace_id, agent_id, session_id)
            if not os.path.exists(secure_path):
                return {"error": f"File not found: {path}"}
            if not path.lower().endswith((".xlsx", ".xlsm")):
                return {"error": "File must have .xlsx or .xlsm extension"}
            if row < 1:
                return {"error": "row must be at least 1"}
            wb = load_workbook(secure_path)
            try:
                if sheet:
                    if sheet not in wb.sheetnames:
                        return {"error": f"Sheet '{sheet}' not found"}
                    ws = wb[sheet]
                else:
                    ws = wb.active
                if ws is None:
                    return {"error": "No active sheet"}
                # Determine column index
                if isinstance(column, int):
                    if column < 1:
                        return {"error": "column index must be at least 1"}
                    col_idx = column
                elif isinstance(column, str):
                    # Find column index by name (from header)
                    headers = [cell.value for cell in ws[1] if cell.value is not None]
                    if column not in headers:
                        return {"error": f"Column '{column}' not found in header"}
                    col_idx = headers.index(column) + 1
                else:
                    return {"error": "column must be str (name) or int (index)"}
                # Update cell
                cell = ws.cell(row=row, column=col_idx, value=value)
                wb.save(secure_path)
                return {
                    "success": True,
                    "path": path,
                    "sheet_name": ws.title,
                    "updated_row": row,
                    "updated_column": get_column_letter(col_idx),
                    "new_value": _convert_cell_value(value),
                }
            finally:
                wb.close()
        except Exception as e:
            return {"error": f"Failed to update cell: {str(e)}"}

    @mcp.tool()
    def excel_delete_row(
        path: str,
        workspace_id: str,
        agent_id: str,
        session_id: str,
        sheet: str | None = None,
        row: int,
    ) -> dict:
        """
        Delete a specific row from an Excel sheet.
        Args:
            path: Path to the Excel file (relative to session sandbox)
            workspace_id: Workspace identifier
            agent_id: Agent identifier
            session_id: Session identifier
            sheet: Sheet name (default: active sheet)
            row: 1-indexed row number to delete (cannot delete header row 1)
        Returns:
            dict with success status and metadata
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            return {
                "error": (
                    "openpyxl not installed. Install with: "
                    "pip install openpyxl or pip install tools[excel]"
                )
            }
        try:
            secure_path = get_secure_path(path, workspace_id, agent_id, session_id)
            if not os.path.exists(secure_path):
                return {"error": f"File not found: {path}"}
            if not path.lower().endswith((".xlsx", ".xlsm")):
                return {"error": "File must have .xlsx or .xlsm extension"}
            if row < 2:
                return {"error": "Cannot delete header row (row 1). Row must be at least 2"}
            wb = load_workbook(secure_path)
            try:
                if sheet:
                    if sheet not in wb.sheetnames:
                        return {"error": f"Sheet '{sheet}' not found"}
                    ws = wb[sheet]
                else:
                    ws = wb.active
                if ws is None:
                    return {"error": "No active sheet"}
                if row > ws.max_row:
                    return {"error": f"Row {row} exceeds max row {ws.max_row}"}
                ws.delete_rows(row)
                wb.save(secure_path)
                return {
                    "success": True,
                    "path": path,
                    "sheet_name": ws.title,
                    "deleted_row": row,
                    "new_row_count": ws.max_row - 1,  # excluding header
                }
            finally:
                wb.close()
        except Exception as e:
            return {"error": f"Failed to delete row: {str(e)}"}

    @mcp.tool()
    def excel_add_column(
        path: str,
        workspace_id: str,
        agent_id: str,
        session_id: str,
        sheet: str | None = None,
        column_name: str,
        default_value: Any = None,
    ) -> dict:
        """
        Add a new column to an Excel sheet.
        Args:
            path: Path to the Excel file (relative to session sandbox)
            workspace_id: Workspace identifier
            agent_id: Agent identifier
            session_id: Session identifier
            sheet: Sheet name (default: active sheet)
            column_name: Name for the new column
            default_value: Default value for all rows in the new column (default: None)
        Returns:
            dict with success status and metadata
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            return {
                "error": (
                    "openpyxl not installed. Install with: "
                    "pip install openpyxl or pip install tools[excel]"
                )
            }
        try:
            secure_path = get_secure_path(path, workspace_id, agent_id, session_id)
            if not os.path.exists(secure_path):
                return {"error": f"File not found: {path}"}
            if not path.lower().endswith((".xlsx", ".xlsm")):
                return {"error": "File must have .xlsx or .xlsm extension"}
            if not column_name:
                return {"error": "column_name cannot be empty"}
            wb = load_workbook(secure_path)
            try:
                if sheet:
                    if sheet not in wb.sheetnames:
                        return {"error": f"Sheet '{sheet}' not found"}
                    ws = wb[sheet]
                else:
                    ws = wb.active
                if ws is None:
                    return {"error": "No active sheet"}
                # Check if column exists
                headers = [str(cell.value) for cell in ws[1] if cell.value is not None]
                if column_name in headers:
                    return {"error": f"Column '{column_name}' already exists"}
                # Add to header
                new_col_idx = ws.max_column + 1
                ws.cell(row=1, column=new_col_idx, value=column_name)
                # Set default value for data rows
                for row_idx in range(2, ws.max_row + 1):
                    ws.cell(row=row_idx, column=new_col_idx, value=default_value)
                wb.save(secure_path)
                return {
                    "success": True,
                    "path": path,
                    "sheet_name": ws.title,
                    "new_column": column_name,
                    "column_count": ws.max_column,
                }
            finally:
                wb.close()
        except Exception as e:
            return {"error": f"Failed to add column: {str(e)}"}

    @mcp.tool()
    def excel_sort(
        path: str,
        workspace_id: str,
        agent_id: str,
        session_id: str,
        sheet: str | None = None,
        sort_by: list[str],
        ascending: bool = True,
    ) -> dict:
        """
        Sort the data in an Excel sheet by specified columns.
        Args:
            path: Path to the Excel file (relative to session sandbox)
            workspace_id: Workspace identifier
            agent_id: Agent identifier
            session_id: Session identifier
            sheet: Sheet name (default: active sheet)
            sort_by: List of column names to sort by (in order of priority)
            ascending: True for ascending sort, False for descending (applies to all columns)
        Returns:
            dict with success status and metadata
        """
        try:
            from openpyxl import load_workbook
            import pandas as pd
        except ImportError:
            return {
                "error": (
                    "openpyxl or pandas not installed. Install with: "
                    "pip install openpyxl pandas or pip install tools[excel]"
                )
            }
        try:
            secure_path = get_secure_path(path, workspace_id, agent_id, session_id)
            if not os.path.exists(secure_path):
                return {"error": f"File not found: {path}"}
            if not path.lower().endswith((".xlsx", ".xlsm")):
                return {"error": "File must have .xlsx or .xlsm extension"}
            if not sort_by:
                return {"error": "sort_by cannot be empty"}
            wb = load_workbook(secure_path, data_only=True)
            try:
                if sheet:
                    if sheet not in wb.sheetnames:
                        return {"error": f"Sheet '{sheet}' not found"}
                    ws = wb[sheet]
                else:
                    ws = wb.active
                if ws is None:
                    return {"error": "No active sheet"}
                # Load data into pandas
                data = ws.iter_rows(values_only=True)
                headers = next(data)
                df = pd.DataFrame(data, columns=headers)
                # Validate sort columns
                for col in sort_by:
                    if col not in df.columns:
                        return {"error": f"Sort column '{col}' not found"}
                # Sort
                df = df.sort_values(by=sort_by, ascending=ascending)
                # Clear sheet (except header)
                ws.delete_rows(2, ws.max_row)
                # Write sorted data
                for row_idx, row in enumerate(df.itertuples(index=False), start=2):
                    for col_idx, value in enumerate(row, start=1):
                        ws.cell(row=row_idx, column=col_idx, value=value)
                wb.save(secure_path)
                return {
                    "success": True,
                    "path": path,
                    "sheet_name": ws.title,
                    "sorted_by": sort_by,
                    "ascending": ascending,
                }
            finally:
                wb.close()
        except Exception as e:
            return {"error": f"Failed to sort sheet: {str(e)}"}

def _convert_cell_value(value: Any) -> Any:
    """Convert Excel cell values to JSON-serializable types."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, (int, float, str, bool)):
        return value
    # For any other type, convert to string
    return str(value)
